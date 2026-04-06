# -*- coding: utf-8 -*-
"""
Batch evaluation for backend (standalone) — mirrors eval_videos_vit_facecrop.py logic.

NOW SUPPORTS:
- Single-model evaluation (như cũ)
- Multi-model ensemble: trung bình xác suất giữa các model cùng kiến trúc

- For REAL videos:
    accuracy = r_frames / n_frames,
    with r_frames = n_frames - f_frames,
    where f_frames = count(p_fake_ensemble >= thr)

- For FAKE videos:
    video_accuracy = c_frames / n_frames,
    where:
        p_fake_ensemble = average_j p_fake_j
        p_method_ensemble = average_j p_method_j
        c_frames = count(p_fake_ensemble >= thr AND pred_method_ensemble == true_method)

    fake_cnt = count(p_fake_ensemble >= thr)
    m_frames = fake_cnt - c_frames

Face detector:
    default: RetinaFace (InsightFace, GPU if available). Fallback: CPU provider.
    option:  --detector_backend mediapipe  (CPU)

Outputs:
    results_real.csv  and  results_fake.csv
"""

import os, csv, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Tuple, Optional

import numpy as np
import torch
import torch.nn as nn
import timm
import cv2
from torchvision import transforms
from tqdm import tqdm


# ── Excel report builder ──────────────────────────────────────────────────────
FAKE_METHODS = ["Audio2Animation", "Deepfakes", "Face2Face", "FaceShifter",
                "FaceSwap", "NeuralTextures", "Video2VideoID"]

def _cell_style(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border

def build_excel_report(fake_csv: str, real_csv: str, out_xlsx: str):
    import csv as _csv

    # ── đọc CSV ──
    def read_csv(path):
        rows = []
        with open(path, newline="", encoding="utf-8") as f:
            reader = _csv.DictReader(f)
            for r in reader:
                rows.append(r)
        return rows

    fake_rows = read_csv(fake_csv) if os.path.exists(fake_csv) else []
    real_rows = read_csv(real_csv) if os.path.exists(real_csv) else []

    # ── styles ──
    hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill    = PatternFill("solid", start_color="2E4057")
    sum_fill    = PatternFill("solid", start_color="E8F4FD")
    avg_fill    = PatternFill("solid", start_color="D4EDDA")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"),  bottom=Side(style="thin"))
    bold_font   = Font(name="Arial", bold=True, size=10)
    norm_font   = Font(name="Arial", size=10)

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════
    # SHEET FAKE
    # ════════════════════════════════════════════════════════════════
    ws_f = wb.active
    ws_f.title = "fake"

    # -- cột data gốc --
    data_cols = ["Videoid","n_frames","f_frames","r_frames","c_frames",
                 "m_frames","duration","threshold","accuracy","method"]
    # -- cột summary (bắt đầu từ cột K = 11) --
    # K: Number of Videos, L: Method/Total, M: (trống), N: (trống),
    # O: Avg Accuracy,  P: >0.9,  Q:(label),  R: 0.7-0.9,  S:(label),
    # T: 0.5-0.7,  U:(label),  V: <0.5,  W:(label),  X: =0

    fake_headers = data_cols + ["Number of Videos","Method/Total","","",
                                "Avg Accuracy",">0.9","",
                                "0.7-0.9","","0.5-0.7","","<0.5","","=0"]
    ws_f.append(fake_headers)
    for c in range(1, len(fake_headers)+1):
        cell = ws_f.cell(row=1, column=c)
        _cell_style(cell, font=hdr_font, fill=hdr_fill, align=center, border=thin)
    ws_f.row_dimensions[1].height = 24
    ws_f.freeze_panes = "A2"

    # -- ghi data rows --
    def to_float(v, default=0.0):
        try: return float(v)
        except: return default

    def to_int(v, default=0):
        try: return int(float(v))
        except: return default

    # group by method
    from collections import defaultdict
    by_method = defaultdict(list)
    for r in fake_rows:
        by_method[r["method"]].append(r)

    data_row = 2  # current Excel row

    # summary accumulators
    all_method_summaries = []   # (method, n_videos, avg_acc, buckets)
    total_videos = 0

    for method in FAKE_METHODS:
        rows_m = by_method.get(method, [])
        start_row = data_row

        for r in rows_m:
            vals = [to_int(r["Videoid"]), to_int(r["n_frames"]),
                    to_int(r["f_frames"]), to_int(r["r_frames"]),
                    to_int(r["c_frames"]), to_int(r["m_frames"]),
                    to_float(r["duration"]), to_float(r["threshold"]),
                    to_float(r["accuracy"]), r["method"]]
            vals += [None]*14
            ws_f.append(vals)
            for c in range(1, 11):
                cell = ws_f.cell(row=data_row, column=c)
                _cell_style(cell, font=norm_font, align=center if c != 1 else center, border=thin)
            data_row += 1

        if not rows_m:
            continue

        # -- summary cho method này --
        accs = [to_float(r["accuracy"]) for r in rows_m]
        n    = len(accs)
        avg  = sum(accs) / n if n else 0
        b90  = sum(1 for a in accs if a >= 0.9)
        b70  = sum(1 for a in accs if 0.7 <= a < 0.9)
        b50  = sum(1 for a in accs if 0.5 <= a < 0.7)
        b0   = sum(1 for a in accs if a < 0.5)
        b00  = sum(1 for a in accs if a == 0.0)
        total_videos += n
        all_method_summaries.append((method, n, avg, b90, b70, b50, b0, b00))

        # ghi summary vào dòng đầu tiên của method
        sr = start_row
        ws_f.cell(row=sr, column=11).value = n          # Number of Videos
        ws_f.cell(row=sr, column=12).value = method     # Method/Total
        ws_f.cell(row=sr, column=15).value = round(avg, 6)  # Avg Accuracy
        ws_f.cell(row=sr, column=16).value = b90
        ws_f.cell(row=sr, column=18).value = b70
        ws_f.cell(row=sr, column=20).value = b50
        ws_f.cell(row=sr, column=22).value = b0
        ws_f.cell(row=sr, column=24).value = b00

        for c in [11, 12, 15, 16, 18, 20, 22, 24]:
            cell = ws_f.cell(row=sr, column=c)
            _cell_style(cell, font=norm_font, fill=sum_fill, align=center, border=thin)

    # -- Average Accuracy row (tổng kết) --
    if all_method_summaries:
        # lấy dòng đầu tiên của method đầu tiên để viết vào (như file gốc)
        # thực ra file gốc viết vào dòng đầu của method cuối cùng
        # nhưng ta viết một dòng riêng
        ws_f.append([None]*10 + [None]*14)
        avg_row = data_row

        all_accs_flat = [to_float(r["accuracy"]) for r in fake_rows]
        n_total  = len(all_accs_flat)
        avg_total = sum(all_accs_flat) / n_total if n_total else 0
        b90t = sum(1 for a in all_accs_flat if a >= 0.9)
        b70t = sum(1 for a in all_accs_flat if 0.7 <= a < 0.9)
        b50t = sum(1 for a in all_accs_flat if 0.5 <= a < 0.7)
        b0t  = sum(1 for a in all_accs_flat if a < 0.5)
        b00t = sum(1 for a in all_accs_flat if a == 0.0)

        ws_f.cell(row=avg_row, column=11).value = n_total
        ws_f.cell(row=avg_row, column=12).value = "Average Accuracy"
        ws_f.cell(row=avg_row, column=15).value = round(avg_total, 6)
        ws_f.cell(row=avg_row, column=16).value = b90t
        ws_f.cell(row=avg_row, column=18).value = b70t
        ws_f.cell(row=avg_row, column=20).value = b50t
        ws_f.cell(row=avg_row, column=22).value = b0t
        ws_f.cell(row=avg_row, column=24).value = b00t

        for c in [11, 12, 15, 16, 18, 20, 22, 24]:
            cell = ws_f.cell(row=avg_row, column=c)
            _cell_style(cell, font=Font(name="Arial", bold=True, size=10),
                        fill=avg_fill, align=center, border=thin)
        data_row += 1

    # column widths fake
    fake_col_w = [10,10,10,10,10,10,10,12,10,18,16,18,6,6,14,8,6,8,6,8,6,8,6,6]
    for i, w in enumerate(fake_col_w, 1):
        ws_f.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════════════
    # SHEET REAL
    # ════════════════════════════════════════════════════════════════
    ws_r = wb.create_sheet("real")

    real_data_cols = ["Videoid","n_frames","f_frames","r_frames",
                      "duration","threshold","accuracy","dataset"]
    real_headers = real_data_cols + ["Number of Videos","Average accuracy","","",
                                     ">0.9","","0.7-0.9","","0.5-0.7","","<0.5","","=0"]
    ws_r.append(real_headers)
    for c in range(1, len(real_headers)+1):
        cell = ws_r.cell(row=1, column=c)
        _cell_style(cell, font=hdr_font, fill=hdr_fill, align=center, border=thin)
    ws_r.row_dimensions[1].height = 24
    ws_r.freeze_panes = "A2"

    by_dataset = defaultdict(list)
    for r in real_rows:
        by_dataset[r["dataset"]].append(r)

    real_row = 2
    DATASET_ORDER = ["Celeb-DF-v2", "Faceforensics", "Tiktok-DF"]
    all_datasets = DATASET_ORDER + [d for d in by_dataset if d not in DATASET_ORDER]

    for ds in all_datasets:
        rows_d = by_dataset.get(ds, [])
        if not rows_d:
            continue
        start_row = real_row

        for r in rows_d:
            vals = [to_int(r["Videoid"]), to_int(r["n_frames"]),
                    to_int(r["f_frames"]), to_int(r["r_frames"]),
                    to_float(r["duration"]), to_float(r["threshold"]),
                    to_float(r["accuracy"]), r["dataset"]]
            vals += [None]*13
            ws_r.append(vals)
            for c in range(1, 9):
                cell = ws_r.cell(row=real_row, column=c)
                _cell_style(cell, font=norm_font, align=center, border=thin)
            real_row += 1

        accs = [to_float(r["accuracy"]) for r in rows_d]
        n    = len(accs)
        avg  = sum(accs) / n if n else 0
        b90  = sum(1 for a in accs if a >= 0.9)
        b70  = sum(1 for a in accs if 0.7 <= a < 0.9)
        b50  = sum(1 for a in accs if 0.5 <= a < 0.7)
        b0   = sum(1 for a in accs if a < 0.5)
        b00  = sum(1 for a in accs if a == 0.0)

        ws_r.cell(row=start_row, column=9).value  = n
        ws_r.cell(row=start_row, column=10).value = round(avg, 6)
        ws_r.cell(row=start_row, column=13).value = b90
        ws_r.cell(row=start_row, column=15).value = b70
        ws_r.cell(row=start_row, column=17).value = b50
        ws_r.cell(row=start_row, column=19).value = b0
        ws_r.cell(row=start_row, column=21).value = b00

        for c in [9, 10, 13, 15, 17, 19, 21]:
            cell = ws_r.cell(row=start_row, column=c)
            _cell_style(cell, font=norm_font, fill=sum_fill, align=center, border=thin)

    # -- Overall real summary --
    if real_rows:
        ws_r.append([None]*8 + [None]*13)
        avg_row_r = real_row
        all_real_accs = [to_float(r["accuracy"]) for r in real_rows]
        n_r   = len(all_real_accs)
        avg_r = sum(all_real_accs) / n_r if n_r else 0
        b90r  = sum(1 for a in all_real_accs if a >= 0.9)
        b70r  = sum(1 for a in all_real_accs if 0.7 <= a < 0.9)
        b50r  = sum(1 for a in all_real_accs if 0.5 <= a < 0.7)
        b0r   = sum(1 for a in all_real_accs if a < 0.5)
        b00r  = sum(1 for a in all_real_accs if a == 0.0)

        ws_r.cell(row=avg_row_r, column=9).value  = n_r
        ws_r.cell(row=avg_row_r, column=10).value = round(avg_r, 6)
        ws_r.cell(row=avg_row_r, column=8).value  = "Overall"
        ws_r.cell(row=avg_row_r, column=13).value = b90r
        ws_r.cell(row=avg_row_r, column=15).value = b70r
        ws_r.cell(row=avg_row_r, column=17).value = b50r
        ws_r.cell(row=avg_row_r, column=19).value = b0r
        ws_r.cell(row=avg_row_r, column=21).value = b00r

        for c in [8, 9, 10, 13, 15, 17, 19, 21]:
            cell = ws_r.cell(row=avg_row_r, column=c)
            _cell_style(cell, font=Font(name="Arial", bold=True, size=10),
                        fill=avg_fill, align=center, border=thin)

    real_col_w = [10,10,10,10,10,12,10,16,16,14,6,6,8,6,8,6,8,6,8,6,6]
    for i, w in enumerate(real_col_w, 1):
        ws_r.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_xlsx)
    print(f"[DONE] Excel report saved -> {out_xlsx}")


# ── Default ensemble weights (tính từ test/val set accuracy từng mô hình) ─────────
# Thứ tự phải khớp với thứ tự --ckpt truyền vào
# bin_weights : trọng số cho p_fake (quyết định nhị phân fake/real)
# met_weights : trọng số cho p_method theo từng loại deepfake
# Normalize về mean=1.0. Đặt None để dùng simple average (không weight).
DEFAULT_WEIGHTS = {
    "model_order": [
        "BEiT-B",
        "ConvNeXt-B",
        "EfficientNet-B5",
        "ResNet-RS152",
        "ViT-B1K",
        "ViT-B21K"
    ],
    "bin_weights": {
        "BEiT-B":          1,
        "ConvNeXt-B":      1,
        "EfficientNet-B5": 1,
        "ResNet-RS152":    1,
        "ViT-B1K":         1,
        "ViT-B21K":        1
    },
    "met_weights": {
        "Audio2Animation": {"BEiT-B": 1, "ConvNeXt-B": 1, "EfficientNet-B5": 1, "ResNet-RS152": 1, "ViT-B1K": 1, "ViT-B21K": 1},
        "Deepfakes":       {"BEiT-B": 1, "ConvNeXt-B": 1, "EfficientNet-B5": 1, "ResNet-RS152": 1, "ViT-B1K": 1, "ViT-B21K": 1},
        "Face2Face":       {"BEiT-B": 1, "ConvNeXt-B": 1, "EfficientNet-B5": 1, "ResNet-RS152": 1, "ViT-B1K": 1, "ViT-B21K": 1},
        "FaceShifter":     {"BEiT-B": 1, "ConvNeXt-B": 1, "EfficientNet-B5": 1, "ResNet-RS152": 1, "ViT-B1K": 1, "ViT-B21K": 1},
        "FaceSwap":        {"BEiT-B": 1, "ConvNeXt-B": 1, "EfficientNet-B5": 1, "ResNet-RS152": 1, "ViT-B1K": 1, "ViT-B21K": 1},
        "NeuralTextures":  {"BEiT-B": 1, "ConvNeXt-B": 1, "EfficientNet-B5": 1, "ResNet-RS152": 1, "ViT-B1K": 1, "ViT-B21K": 1},
        "Video2VideoID":   {"BEiT-B": 1, "ConvNeXt-B": 1, "EfficientNet-B5": 1, "ResNet-RS152": 1, "ViT-B1K": 1, "ViT-B21K": 1}
    }
}


# ── Confusion Matrix Excel Builder ───────────────────────────────────────────
FAKE_METHODS = ["Audio2Animation", "Deepfakes", "Face2Face", "FaceShifter",
                "FaceSwap", "NeuralTextures", "Video2VideoID"]
REAL_DATASETS = ["Celeb-DF-v2", "Faceforensics", "Tiktok-DF"]

METHOD_THEME = {
    "Audio2Animation": 7,
    "Deepfakes":       5,
    "FaceShifter":     2,
    "FaceSwap":        9,
    "NeuralTextures":  8,
}
METHOD_RGB = {
    "Face2Face":            "FF0000",
    "Video2VideoID":        "FFFF00",
    "Real (Celeb-DF-v2)":   "00B050",
    "Real (Faceforensics)": "00B050",
    "Real (Tiktok-DF)":     "00B050",
}
HEADER_BG   = "D9D9D9"   # White Darker 15%
OFF_FAKE_BG = "FFE5E5"   # nhầm fake→fake
OFF_REAL_BG = "EADCF8"   # nhầm real→fake (tím nhạt)
REAL_WRONG  = "FF0000"   # fake→real

def _acc_color(v):
    if v >= 90: return "00B0F0"
    if v >= 70: return "92D050"
    if v >= 50: return "FFFF00"
    if v >  0:  return "FFC000"
    return "FF0000"

def _fill_for(label):
    from openpyxl.styles import PatternFill
    from openpyxl.styles.colors import Color
    if label in METHOD_THEME:
        return PatternFill(patternType="solid", fgColor=Color(theme=METHOD_THEME[label], type="theme"))
    if label in METHOD_RGB:
        return PatternFill("solid", start_color=METHOD_RGB[label])
    return None

def _s(ws, r, c, val, fill=None, rgb=None, bold=False, align="center",
       fs=11, border=True):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    cell = ws.cell(row=r, column=c)
    cell.value = val
    if fill:   cell.fill = fill
    elif rgb:  cell.fill = PatternFill("solid", start_color=("FF"+rgb if len(rgb)==6 else rgb))
    cell.font      = Font(bold=bold, color="000000", size=fs, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        s = Side(style="thin")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
    return cell

def build_confusion_matrix_excel(fake_csv, real_csv, out_xlsx, model_name="Model"):
    import csv as _csv, os
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    def read_csv(path):
        if not path or not os.path.exists(path): return []
        with open(path, newline="", encoding="utf-8") as f:
            return list(_csv.DictReader(f))

    def ti(v):
        try: return int(float(v))
        except: return 0
    def tf(v):
        try: return float(v)
        except: return 0.0

    fake_rows = read_csv(fake_csv)
    real_rows = read_csv(real_csv)
    sample    = fake_rows[0] if fake_rows else {}
    has_pred  = any(k.startswith("pred_") for k in sample.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = model_name[:31]

    # ── Column widths ──
    for col, w in [("A",20),("B",10),("C",13),("D",14),("E",14),("F",14),
                   ("G",14),("H",14),("I",14),("J",14),("K",14),("L",10),
                   ("M",18),("N",14),("O",16),("P",16),("Q",14),("R",14),
                   ("S",10),("T",10),("U",10),("V",10),("W",10)]:
        ws.column_dimensions[col].width = w

    # ── ROW 1-2: Headers ──
    # Các cột merge rows 1-2
    for col, label in [(1,"Deepfake Types"),(2,"VideoID"),(3,"Total Frames"),
                       (12,"Duration"),(13,"Threshold"),
                       (14,"Accuracy (%)"),(15,"Avg Cls Acc (%)"),(16,"Median Acc (%)"),(17,"Std (method)")]:
        _s(ws,1,col,label,bold=True,rgb=HEADER_BG)
        ws.merge_cells(start_row=1,start_column=col,end_row=2,end_column=col)
    # K: Real header — xanh lá thuần
    ws.merge_cells(start_row=1,start_column=11,end_row=2,end_column=11)

    # D1:J1 merge (Fake header)
    _s(ws,1,4,"Fake",bold=True,rgb="FFBABA")
    ws.merge_cells("D1:J1")

    # P1:T1 merge (legend header — D9D9D9)
    _s(ws,1,18,"",rgb=HEADER_BG)
    ws.merge_cells("R1:V1")

    # K1 Real — set SAU merge để không bị ghi đè bởi D9D9D9
    _s(ws,1,11,"Real",bold=True,rgb="00B050")

    # ROW 2: sub-headers 7 fake methods
    for ci, mname in enumerate(FAKE_METHODS, 4):
        _s(ws,2,ci,mname,bold=True,fill=_fill_for(mname))

    # ROW 2: legend P-T
    for ci,(lbl,bg) in enumerate([(">=90%","00B0F0"),("90%-70%","92D050"),
                                   ("70%-50%","FFFF00"),(">50%","FFC000"),
                                   ("0%","FF0000")], 18):
        _s(ws,2,ci,lbl,bold=True,rgb=bg)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # ── Group data ──
    fake_by_method = defaultdict(list)
    for r in fake_rows: fake_by_method[r["method"]].append(r)

    real_by_ds = defaultdict(list)
    for r in real_rows: real_by_ds[f"Real ({r.get('dataset','Unknown')})"].append(r)

    current_row = 3

    def write_group(label, videos, is_real=False):
        nonlocal current_row
        gfill    = _fill_for(label)
        true_idx = FAKE_METHODS.index(label) if label in FAKE_METHODS else -1
        n        = len(videos)

        for idx, r in enumerate(videos):
            vid_id    = r.get("Videoid", r.get("VideoID",""))
            n_frames  = ti(r.get("n_frames",0))
            duration  = round(tf(r.get("duration",0)),2)
            threshold = tf(r.get("threshold",0))
            acc_pct   = round(tf(r.get("accuracy",0))*100, 1)
            is_last   = (idx == n - 1)

            # A: label
            _s(ws, current_row, 1, label, fill=gfill)
            # B,C
            _s(ws, current_row, 2, ti(vid_id) if str(vid_id).isdigit() else vid_id)
            _s(ws, current_row, 3, n_frames)

            # D-J: confusion columns
            for ci, mname in enumerate(FAKE_METHODS, 4):
                col_idx = ci - 4
                # Với real rows: pred_<method> là số frames bị nhầm sang loại đó
                # has_pred kiểm tra cả fake và real CSV
                has_pred_real = any(f"pred_{m}" in r for m in FAKE_METHODS)
                val = ti(r.get(f"pred_{mname}", 0)) if has_pred_real else 0

                if is_real:
                    # Real rows: nhầm sang fake — màu tím nhạt
                    bg = OFF_REAL_BG if val > 0 else None
                    _s(ws, current_row, ci, val, rgb=bg)
                else:
                    if col_idx == true_idx:
                        # Diagonal
                        if not has_pred: val = ti(r.get("c_frames", 0))
                        _s(ws, current_row, ci, val, fill=gfill)
                    else:
                        # Off-diagonal fake→fake — màu hồng nhạt
                        bg = OFF_FAKE_BG if val > 0 else None
                        _s(ws, current_row, ci, val, rgb=bg)

            # K: Real column
            if is_real:
                r_frames = ti(r.get("r_frames", 0))
                _s(ws, current_row, 11, r_frames, fill=gfill)
            else:
                real_val = ti(r.get("pred_Real", 0)) if has_pred else 0
                bg_k = REAL_WRONG if real_val > 0 else None
                _s(ws, current_row, 11, real_val, rgb=bg_k)

            # L,M,N
            _s(ws, current_row, 12, duration)
            _s(ws, current_row, 13, threshold)
            _s(ws, current_row, 14, acc_pct, rgb=_acc_color(acc_pct))

            # O-R + legend S-W: chỉ hàng CUỐI
            if is_last:
                import statistics as _stats
                all_cls  = [tf(v.get("accuracy",0))*100 for v in videos]
                avg_cls = round(sum(all_cls)/len(all_cls),2) if all_cls else 0
                med_cls = round(_stats.median(all_cls),2) if all_cls else 0
                std_cls = round(_stats.stdev(all_cls),2) if len(all_cls)>1 else 0.0
                _s(ws, current_row, 15, avg_cls, fill=gfill)   # O: Avg Cls
                _s(ws, current_row, 16, med_cls, fill=gfill)   # P: Median
                _s(ws, current_row, 17, std_cls, fill=gfill)   # Q: Std (method)
                b = [sum(1 for a in all_cls if a>=90),
                     sum(1 for a in all_cls if 70<=a<90),
                     sum(1 for a in all_cls if 50<=a<70),
                     sum(1 for a in all_cls if 0<a<50),
                     sum(1 for a in all_cls if a==0)]
                for ci,(bv,bg) in enumerate(zip(b,["00B0F0","92D050","FFFF00","FFC000","FF0000"]),18):
                    _s(ws, current_row, ci, bv if bv>0 else 0, rgb=bg)

            ws.row_dimensions[current_row].height = 18
            current_row += 1

    for method in FAKE_METHODS:
        write_group(method, fake_by_method.get(method,[]), is_real=False)
    for ds in REAL_DATASETS:
        write_group(f"Real ({ds})", real_by_ds.get(f"Real ({ds})",[]), is_real=True)

    # ── Summary ──
    sr = current_row + 2
    ws.merge_cells(f"E{sr}:I{sr}")
    _s(ws,sr,5,"Summary",bold=True,rgb=HEADER_BG)
    # Border E:I
    from openpyxl.styles import Border as _Brd, Side as _Sd
    _ts = _Sd(style="thin")
    for _c in range(5, 10):
        ws.cell(row=sr,column=_c).border = _Brd(
            left  =_ts if _c==5 else None,
            right =_ts if _c==9 else None,
            top   =_ts, bottom=_ts)

    # Hàng header
    _s(ws,sr+1,5,"Overall Accuracy (%)",bold=True,rgb="00B0F0")
    _s(ws,sr+1,6,"Fake Cls (%)",  bold=True,rgb="FF0000")
    _s(ws,sr+1,7,"Real (%)",      bold=True,rgb="00B050")
    _s(ws,sr+1,8,"Median (%)",    bold=True,rgb=HEADER_BG)
    _s(ws,sr+1,9,"Std Model",     bold=True,rgb=HEADER_BG)

    # Tính
    import statistics as _stats
    all_fake = [tf(r.get("accuracy",0))*100 for r in fake_rows]
    all_real = [tf(r.get("accuracy",0))*100 for r in real_rows]
    fake_avg = round(sum(all_fake)/len(all_fake),2)   if all_fake else 0
    real_avg = round(sum(all_real)/len(all_real),2)   if all_real else 0
    overall  = round((fake_avg+real_avg)/2,2)
    med_all  = round(_stats.median(all_fake),2)       if all_fake else 0

    # Avg per method → std liên loại
    from collections import defaultdict as _dd
    by_m = _dd(list)
    for r in fake_rows: by_m[r["method"]].append(tf(r.get("accuracy",0))*100)
    per_method_avgs = [sum(v)/len(v) for v in by_m.values() if v]
    std_inter = round(_stats.stdev(per_method_avgs),2) if len(per_method_avgs)>1 else 0.0

    _s(ws,sr+2,5, overall,   rgb="00B0F0",bold=True)
    _s(ws,sr+2,6, fake_avg,  rgb="FFBABA",bold=True)
    _s(ws,sr+2,7, real_avg,  rgb="00B050",bold=True)
    _s(ws,sr+2,8, med_all,   rgb=HEADER_BG)
    _s(ws,sr+2,9, std_inter, rgb=HEADER_BG)

    # ── Chú thích ──
    note_row = sr + 5
    _s(ws, note_row, 5, "Chú thích các cột", bold=True, rgb=HEADER_BG)
    ws.merge_cells(start_row=note_row, start_column=5, end_row=note_row, end_column=10)

    NOTES = [
        ("Deepfake Types",       "Loại kỹ thuật (A2A, DF,...) hoặc nhóm Real."),
        ("VideoID",              "Mã định danh riêng của từng video test."),
        ("Total Frames",         "Tổng số khung hình trong video."),
        ("Duration",             "Thời lượng video (giây)."),
        ("Threshold",            "Ngưỡng xác suất để phân loại."),
        ("Accuracy (%)",         "Tỉ lệ đoán đúng trên từng video cụ thể."),
        # Chỉ số mức độ Loại (Method level)
        ("Avg Cls Acc (%)",      "Tỉ lệ nhận diện đúng loại Deepfake (Trung bình của loại đó)."),
        ("Median Acc (%)",       "Trung vị của loại Deepfake đó (càng gần Avg Cls Acc càng tốt)."),
        ("Std (method)",         "Độ lệch chuẩn - đo sự ổn định trên loại Deepfake đó."),
        # Chỉ số mức độ Mô hình (Model level)
        ("Overall Accuracy (%)", "Độ chính xác tổng thể (Trung bình của cả Deepfake và Real)."),
        ("Fake Cls (%)",         "Khả năng nhận diện đúng loại trên toàn bộ video Deepfake."),
        ("Real (%)",             "Khả năng nhận diện đúng video Người thật không có Deepfake."),
        ("Median (%)",           "Trung vị độ chính xác của toàn bộ Mô hình."),
        ("Std Model",            "Độ lệch chuẩn giữa các loại Deepfake - khả năng tổng quát hóa của mô hình đối với các kiểu Deepfake khác nhau."),
    ]
    from openpyxl.styles import Alignment as _Aln, Border as _Brd, Side as _Sd
    data_start = note_row + 1
    for idx_n, (col_name, desc) in enumerate(NOTES):
        r1 = data_start + idx_n * 2
        r2 = r1 + 1
        # Col E: tên — merge 2 hàng dọc
        _s(ws, r1, 5, col_name, bold=True, rgb=HEADER_BG, align="center")
        ws.merge_cells(start_row=r1, start_column=5, end_row=r2, end_column=5)
        ws.cell(r1, 5).alignment = _Aln(horizontal="center", vertical="center", wrap_text=True)
        # Col F:J: mô tả — merge ngang+dọc 2 hàng
        _s(ws, r1, 6, desc, bold=False, rgb=None, align="left")
        ws.merge_cells(start_row=r1, start_column=6, end_row=r2, end_column=10)
        ws.cell(r1, 6).alignment = _Aln(horizontal="left", vertical="center", wrap_text=True)
        # Border ô r2 col E (phần dưới merge dọc)
        _t = _Sd(style="thin")
        ws.cell(r2, 5).border = _Brd(left=_t, right=_t, bottom=_t)
        ws.row_dimensions[r1].height = 18
        ws.row_dimensions[r2].height = 18
    ws.row_dimensions[note_row].height = 20

    wb.save(out_xlsx)
    print(f"[DONE] Confusion Matrix -> {out_xlsx}")


IMAGENET_MEAN = (0.485, 0.456, 0.406)
IMAGENET_STD  = (0.229, 0.224, 0.225)
VIDEO_EXTS = {".mp4", ".avi", ".mov", ".mkv", ".webm", ".m4v"}

# ---------------- Model khớp train_vit.py / train_spatial.py ----------------
class MultiHeadViT(nn.Module):
    """
    Wrapper timm backbone + 5 head:
      - head_bin:   2 lớp (real/fake)
      - head_met:   num_methods
      - head_face:  num_face_classes
      - head_head:  num_head_classes
      - head_full:  num_full_classes

    Dùng chung được cho ViT, ConvNeXt, Swin... miễn timm hỗ trợ model_name.
    """
    def __init__(self, model_name: str, img_size: int,
                 num_methods: int, num_face_classes: int, num_head_classes: int, num_full_classes: int,
                 drop_rate: float=0.0, drop_path_rate: float=0.0):
        super().__init__()

        backbone_kwargs = dict(
            pretrained=False,
            num_classes=0,
            drop_rate=drop_rate,
            drop_path_rate=drop_path_rate,
        )

        # Một số backbone (ViT, Swin, BEiT, ...) hỗ trợ img_size, ConvNeXt thì không
        if any(k in model_name.lower() for k in ["vit", "swin", "beit"]):
            backbone_kwargs["img_size"] = img_size

        self.backbone = timm.create_model(
            model_name,
            **backbone_kwargs,
        )
        feat = self.backbone.num_features

        def head(n):
            return nn.Sequential(
                nn.Dropout(p=drop_rate if drop_rate > 0 else 0.0),
                nn.Linear(feat, n)
            )

        self.head_bin  = head(2)
        self.head_met  = head(num_methods)
        self.head_face = head(max(1, num_face_classes))
        self.head_head = head(max(1, num_head_classes))
        self.head_full = head(max(1, num_full_classes))

    def forward(self, x):
        f = self.backbone(x)
        return self.head_bin(f), self.head_met(f), self.head_face(f), self.head_head(f), self.head_full(f)


def _infer_head_sizes_from_ckpt_state(ckpt_model_state: Dict[str, torch.Tensor]) -> Dict[str, int]:
    sizes = {"num_methods": 0, "num_face_classes": 1, "num_head_classes": 1, "num_full_classes": 1}
    if "head_met.1.weight"  in ckpt_model_state: sizes["num_methods"]      = ckpt_model_state["head_met.1.weight"].shape[0]
    if "head_face.1.weight" in ckpt_model_state: sizes["num_face_classes"] = ckpt_model_state["head_face.1.weight"].shape[0]
    if "head_head.1.weight" in ckpt_model_state: sizes["num_head_classes"] = ckpt_model_state["head_head.1.weight"].shape[0]
    if "head_full.1.weight" in ckpt_model_state: sizes["num_full_classes"] = ckpt_model_state["head_full.1.weight"].shape[0]
    return sizes


def _filter_state_dict_by_shape(dst_state: Dict[str, torch.Tensor], src_state: Dict[str, torch.Tensor]) -> Dict[str, torch.Tensor]:
    return {k: v for k, v in src_state.items() if k in dst_state and dst_state[k].shape == v.shape}


def load_checkpoint_build_model(ckpt_path: str, device: torch.device,
                                img_size_arg: Optional[int],
                                model_name_arg: Optional[str]) -> Tuple[nn.Module, Dict]:
    ckpt = torch.load(ckpt_path, map_location="cpu")
    meta = ckpt.get("meta", {})
    model_state = ckpt.get("model", {})
    if not model_state:
        raise RuntimeError(f"Checkpoint thiếu key 'model': {ckpt_path}")
    if "best_thr" not in meta and "best_thr" in ckpt:
        meta["best_thr"] = ckpt["best_thr"]
    head_sizes = _infer_head_sizes_from_ckpt_state(model_state)
    num_methods       = head_sizes["num_methods"] or len(meta.get("method_names", [])) or 7
    num_face_classes  = head_sizes["num_face_classes"]
    num_head_classes  = head_sizes["num_head_classes"]
    num_full_classes  = head_sizes["num_full_classes"]

    model_name = model_name_arg or meta.get("backbone_model") or meta.get("model_name", "vit_base_patch16_384")
    img_size   = img_size_arg  or meta.get("img_size", 384)

    model = MultiHeadViT(model_name, img_size, num_methods, num_face_classes, num_head_classes, num_full_classes).to(device)
    dst = model.state_dict()
    if "ema" in ckpt and ckpt["ema"]:
        ema_state = ckpt["ema"]
        dst.update(_filter_state_dict_by_shape(dst, ema_state))
        dst.update(_filter_state_dict_by_shape(dst, model_state))
        model.load_state_dict(dst, strict=False)
    else:
        dst.update(_filter_state_dict_by_shape(dst, model_state))
        model.load_state_dict(dst, strict=False)
    model.eval()

    if not meta.get("method_names"):
        meta["method_names"] = [f"method_{i}" for i in range(num_methods)]
    if "img_size" not in meta:   meta["img_size"]   = img_size
    if "model_name" not in meta: meta["model_name"] = model_name
    if "best_thr" not in meta:   meta["best_thr"]   = 0.5
    return model, meta

# ---------------- Face detection backends ----------------
class RetinaFaceLargest:
    def __init__(self, device: torch.device, det_thresh: float = 0.5, det_size: int = 640):
        from insightface.app import FaceAnalysis
        if device.type == "cuda":
            providers = ["CUDAExecutionProvider", "CPUExecutionProvider"]; ctx_id = 0
        else:
            providers = ["CPUExecutionProvider"]; ctx_id = -1
        self.app = FaceAnalysis(name="buffalo_l", providers=providers)
        self.app.prepare(ctx_id=ctx_id, det_size=(det_size, det_size))
        self.det_thresh = det_thresh

    def detect_largest(self, bgr) -> Optional[Tuple[int,int,int,int]]:
        faces = self.app.get(bgr)
        faces = [f for f in faces if getattr(f, "det_score", 1.0) >= self.det_thresh]
        if not faces: return None
        def area(face): x1,y1,x2,y2 = face.bbox; return (x2-x1)*(y2-y1)
        best = max(faces, key=area)
        x1,y1,x2,y2 = [int(round(v)) for v in best.bbox]
        return x1,y1,x2,y2


def mp_detect_all(bgr) -> List[Tuple[int,int,int,int,float]]:
    import mediapipe as mp
    H,W = bgr.shape[:2]
    det = mp.solutions.face_detection.FaceDetection(model_selection=1, min_detection_confidence=0.5)
    res = det.process(bgr[:, :, ::-1])
    out=[]
    if res.detections:
        for d in res.detections:
            score = d.score[0] if d.score else 0.0
            box = d.location_data.relative_bounding_box
            x = int(round(box.xmin * W)); y = int(round(box.ymin * H))
            w = int(round(box.width * W)); h = int(round(box.height * H))
            x1 = max(0, x); y1 = max(0, y)
            x2 = min(W, x + max(1,w)); y2 = min(H, y + max(1,h))
            if x2 > x1 and y2 > y1: out.append((x1,y1,x2,y2,float(score)))
    return out


def square_crop_from_bbox(bgr, bbox, scale: float=1.10):
    H,W = bgr.shape[:2]
    x1,y1,x2,y2 = bbox
    cx = (x1+x2)/2.0; cy=(y1+y2)/2.0
    side = max(x2-x1, y2-y1) * scale
    nx1 = max(0, int(round(cx-side/2))); ny1 = max(0, int(round(cy-side/2)))
    nx2 = min(W, int(round(cx+side/2))); ny2 = min(H, int(round(cy+side/2)))
    if nx2 <= nx1 or ny2 <= ny1: return None
    return bgr[ny1:ny2, nx1:nx2]


def build_eval_transform(img_size: int):
    return transforms.Compose([
        transforms.ToPILImage(),
        transforms.Resize((img_size, img_size)),
        transforms.ToTensor(),
        transforms.Normalize(IMAGENET_MEAN, IMAGENET_STD),
    ])

# ---------------- Video batching ----------------
def iter_video_batches(video_path: str, batch_size: int):
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened(): raise RuntimeError(f"Cannot open video: {video_path}")
    fps = cap.get(cv2.CAP_PROP_FPS) or 0.0
    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    duration = float(total / fps) if fps > 0 else 0.0

    batches, buf = [], []
    idx = 0
    while True:
        ok, frame_bgr = cap.read()
        if not ok: break
        buf.append((idx, frame_bgr))
        if len(buf) == batch_size:
            batches.append(buf); buf = []
        idx += 1
    if buf: batches.append(buf)
    cap.release()
    return total, duration, batches


def preprocess_face_batch(batch, detector_backend: str, device: torch.device,
                          tx, img_size: int, bbox_scale: float, retina_det: Optional[RetinaFaceLargest]):
    imgs = []; ok_flags=[]
    for _, bgr in batch:
        if detector_backend == "mediapipe":
            dets = mp_detect_all(bgr)
            if dets:
                dets.sort(key=lambda t: t[4], reverse=True)
                x1,y1,x2,y2,_ = dets[0]
                crop = square_crop_from_bbox(bgr, (x1,y1,x2,y2), scale=bbox_scale)
            else:
                crop = None
        else:
            bb = retina_det.detect_largest(bgr) if retina_det is not None else None
            crop = square_crop_from_bbox(bgr, bb, scale=bbox_scale) if bb is not None else None

        if crop is None or crop.size == 0:
            ok_flags.append(False)
            imgs.append(torch.zeros(3, img_size, img_size))
        else:
            imgs.append(tx(crop))
            ok_flags.append(True)
    return torch.stack(imgs, 0), ok_flags

# ---------------- REAL eval ----------------
@torch.no_grad()
def eval_real_dir(models: List[nn.Module], device: torch.device, img_size: int,
                  in_root: Path, out_csv: str,
                  thr: float, batch_size: int, detector_backend: str,
                  bbox_scale: float, fake_index: int,
                  retina_det: Optional[RetinaFaceLargest],
                  model_names: Optional[List[str]] = None,
                  bin_weights: Optional[List[float]] = None,
                  limit_videos: int = 0,
                  method_names: Optional[List[str]] = None):
    """
    Nếu len(models) == 1: dùng đúng model đó.
    Nếu len(models)  > 1: ensemble bằng weighted average p_fake giữa các model.
      - bin_weights: trọng số cho p_fake của từng model (None = simple average)
    """
    tx = build_eval_transform(img_size)
    rows = []
    for ds in [p for p in sorted(in_root.glob("*")) if p.is_dir()]:
        videos = [v for v in sorted(ds.rglob("*")) if v.suffix.lower() in VIDEO_EXTS]
        if limit_videos > 0:
            videos = videos[:limit_videos]
        for v in tqdm(videos, desc=f"[REAL] {ds.name}", unit="vid"):
            try:
                n_frames, duration, batches = iter_video_batches(str(v), batch_size)
                if n_frames == 0:
                    rows.append([v.stem, 0, 0, 0, 0.0, thr, 0.0, ds.name]); continue
                fake_cnt = 0
                # đếm frame real bị nhầm sang từng fake method
                pred_counts_real = {n: 0 for n in method_names} if method_names else {}
                for b in batches:
                    x, ok = preprocess_face_batch(b, detector_backend, device, tx, img_size, bbox_scale, retina_det)
                    x = x.to(device, non_blocking=True)

                    if len(models) == 1:
                        log_bin, log_met, *_ = models[0](x)
                        prob_bin = torch.softmax(log_bin, dim=1)
                        prob_met = torch.softmax(log_met, dim=1)
                    else:
                        prob_bins = []; prob_mets = []
                        for m in models:
                            log_bin, log_met, *_ = m(x)
                            prob_bins.append(torch.softmax(log_bin, dim=1))
                            prob_mets.append(torch.softmax(log_met, dim=1))
                        stacked = torch.stack(prob_bins, dim=0)
                        if bin_weights is not None:
                            w = torch.tensor(bin_weights, dtype=stacked.dtype, device=stacked.device)
                            w = w / w.sum()
                            prob_bin = (stacked * w.view(-1, 1, 1)).sum(dim=0)
                        else:
                            prob_bin = stacked.mean(dim=0)
                        prob_met = torch.stack(prob_mets, dim=0).mean(dim=0)

                    probs  = prob_bin[:, fake_index]
                    m_pred = prob_met.argmax(1)

                    for i, o in enumerate(ok):
                        if not o: continue
                        if probs[i].item() >= thr:
                            fake_cnt += 1
                            # lưu pred method của frame bị nhầm
                            pred_idx = m_pred[i].item()
                            if method_names and pred_idx < len(method_names):
                                pred_counts_real[method_names[pred_idx]] += 1

                real_cnt = n_frames - fake_cnt
                acc = real_cnt / max(1, n_frames)
                pred_cols = [pred_counts_real.get(n,0) for n in (method_names or [])]
                rows.append([v.stem, n_frames, fake_cnt, real_cnt,
                             round(duration,3), thr, round(acc,4), ds.name] + pred_cols)
            except Exception as e:
                n_mn = len(method_names) if method_names else 0
                rows.append([v.stem, 0, 0, 0, 0.0, thr, 0.0,
                             f"{ds.name} (ERROR: {e})"] + [0]*n_mn)
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        pred_headers_real = [f"pred_{n}" for n in (method_names or [])]
        w.writerow(["Videoid","n_frames","f_frames","r_frames",
                    "duration","threshold","accuracy","dataset"] + pred_headers_real)
        w.writerows(rows)

# ---------------- FAKE eval ----------------
@torch.no_grad()
def eval_fake_dir(models: List[nn.Module], device: torch.device, img_size: int,
                  in_root: Path, out_csv: str,
                  thr: float, batch_size: int, detector_backend: str,
                  bbox_scale: float, method_names: List[str], fake_index: int,
                  retina_det: Optional[RetinaFaceLargest],
                  model_names: Optional[List[str]] = None,
                  bin_weights: Optional[List[float]] = None,
                  met_weights: Optional[Dict[str, List[float]]] = None,
                  limit_videos: int = 0):
    """
    Ensemble:
      - p_fake_ensemble = weighted_avg_j softmax(log_bin_j)   [bin_weights]
      - p_met_ensemble  = weighted_avg_j softmax(log_met_j)   [met_weights per method]
      Nếu weights=None thì dùng simple average.
    """
    tx = build_eval_transform(img_size)
    rows = []
    name2idx = {n: i for i, n in enumerate(method_names)}
    for mdir in [p for p in sorted(in_root.glob("*")) if p.is_dir()]:
        mname = mdir.name
        vids = [v for v in sorted(mdir.rglob("*")) if v.suffix.lower() in VIDEO_EXTS]
        if limit_videos > 0:
            vids = vids[:limit_videos]
        for v in tqdm(vids, desc=f"[FAKE] {mname}", unit="vid"):
            try:
                n_frames, duration, batches = iter_video_batches(str(v), batch_size)
                if n_frames == 0:
                    rows.append([v.stem, 0, 0, 0, 0, 0, 0.0, thr, 0.0, mname] + [0]*len(method_names)); continue
                fake_cnt = 0; correct_m = 0
                true_idx = name2idx.get(mname, None)
                # đếm từng pred_method để build confusion matrix
                pred_counts = {n: 0 for n in method_names}
                real_as_real = 0  # frame fake bị nhầm là Real
                for b in batches:
                    x, ok = preprocess_face_batch(b, detector_backend, device, tx, img_size, bbox_scale, retina_det)
                    x = x.to(device, non_blocking=True)

                    if len(models) == 1:
                        log_bin, log_met, *_ = models[0](x)
                        prob_bin = torch.softmax(log_bin, dim=1)
                        prob_met = torch.softmax(log_met, dim=1)
                    else:
                        prob_bins = []
                        prob_mets = []
                        for m in models:
                            log_bin, log_met, *_ = m(x)
                            prob_bins.append(torch.softmax(log_bin, dim=1))
                            prob_mets.append(torch.softmax(log_met, dim=1))
                        stacked_bin = torch.stack(prob_bins, dim=0)  # [N, B, 2]
                        stacked_met = torch.stack(prob_mets, dim=0)  # [N, B, M]
                        # bin_weights: trọng số nhị phân
                        if bin_weights is not None:
                            wb = torch.tensor(bin_weights, dtype=stacked_bin.dtype, device=stacked_bin.device)
                            wb = wb / wb.sum()
                            prob_bin = (stacked_bin * wb.view(-1, 1, 1)).sum(dim=0)
                        else:
                            prob_bin = stacked_bin.mean(dim=0)
                        # met_weights: trọng số per method cho mdir hiện tại
                        if met_weights is not None and mname in met_weights:
                            wm = torch.tensor(met_weights[mname], dtype=stacked_met.dtype, device=stacked_met.device)
                            wm = wm / wm.sum()
                            prob_met = (stacked_met * wm.view(-1, 1, 1)).sum(dim=0)
                        else:
                            prob_met = stacked_met.mean(dim=0)

                    probs  = prob_bin[:, fake_index]
                    m_pred = prob_met.argmax(1)

                    for i, o in enumerate(ok):
                        if not o: continue
                        p_fake = probs[i].item() >= thr
                        if p_fake:
                            fake_cnt += 1
                            pred_idx = m_pred[i].item()
                            if true_idx is not None and pred_idx == true_idx:
                                correct_m += 1
                            # đếm frame bị nhầm sang method nào
                            if pred_idx < len(method_names):
                                pred_counts[method_names[pred_idx]] += 1
                        else:
                            real_as_real += 1  # frame fake bị predict là Real

                real_cnt = n_frames - fake_cnt
                c_frames = correct_m
                m_frames = max(0, fake_cnt - c_frames)
                acc = c_frames / max(1, n_frames)
                # pred_counts theo thứ tự method_names + real_as_real
                pred_cols = [pred_counts.get(n, 0) for n in method_names] + [real_as_real]
                rows.append([v.stem, n_frames, fake_cnt, real_cnt, c_frames, m_frames,
                             round(duration,3), thr, round(acc,4), mname] + pred_cols)
            except Exception as e:
                rows.append([v.stem, 0, 0, 0, 0, 0, 0.0, thr, 0.0, f"{mname} (ERROR: {e})"] + [0]*(len(method_names)+1))
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        pred_headers = [f"pred_{n}" for n in method_names] + ["pred_Real"]
        w.writerow(["Videoid","n_frames","f_frames","r_frames","c_frames","m_frames",
                    "duration","threshold","accuracy","method"] + pred_headers)
        w.writerows(rows)

# ---------------- Main ----------------
def main():
    ap = argparse.ArgumentParser("Backend batch evaluator (mirror eval_videos_vit_facecrop.py) — now with ensemble support")
    ap.add_argument("--root", type=str, required=True, help="data/videos_test (có real/ và fake/)")
    ap.add_argument("--ckpt", type=str, required=True, nargs="+",
                    help="1 hoặc nhiều detector_best.pt (ensemble nếu >=2)")
    ap.add_argument("--device", type=str, default="cuda")
    ap.add_argument("--img_size", type=int, default=0, help="0 = lấy từ ckpt.meta.img_size (ckpt đầu tiên)")
    ap.add_argument("--model_name", type=str, default="", help="rỗng = lấy từ ckpt.meta.model_name/backbone_model")
    ap.add_argument("--batch", type=int, default=64, help="frame batch size")
    ap.add_argument("--threshold", type=float, default=-1.0,
                    help="<0 = 1 model: dùng ckpt.meta.best_thr; >=2 model: dùng average(best_thr) (khuyến nghị: tự tune ensemble thr và truyền vào)")
    ap.add_argument("--bbox_scale", type=float, default=1.10, help="nới ô vuông quanh bbox (1.0 = tight)")
    ap.add_argument("--det_thr", type=float, default=0.5, help="ngưỡng RetinaFace")
    ap.add_argument("--fake_index", type=int, default=0, help="chỉ số lớp FAKE trong head nhị phân (0 hoặc 1)")
    ap.add_argument("--detector_backend", type=str, default="retinaface", choices=["retinaface","mediapipe"])
    ap.add_argument("--out_real", type=str, default="results_real.csv")
    ap.add_argument("--out_fake", type=str, default="results_fake.csv")
    ap.add_argument("--skip_real", action="store_true", help="Bỏ qua test real/")
    ap.add_argument("--skip_fake", action="store_true", help="Bỏ qua test fake/")
    ap.add_argument("--limit_videos", type=int, default=0,
                    help="Giới hạn số video mỗi loại (0 = tất cả). Dùng 1 để test nhanh.")
    ap.add_argument("--out_xlsx", type=str, default="", help="Xuất Excel report (rỗng = không xuất)")
    ap.add_argument("--cm_model_name", type=str, default="Model",
                    help="Tên mô hình hiển thị trong sheet confusion matrix")
    ap.add_argument("--use_weights", action="store_true",
                    help="Dùng DEFAULT_WEIGHTS (bật cả bin_weights + met_weights)")
    ap.add_argument("--use_bin_weights", action="store_true",
                    help="Chỉ dùng bin_weights (trọng số nhị phân fake/real)")
    ap.add_argument("--use_met_weights", action="store_true",
                    help="Chỉ dùng met_weights (trọng số per method)")
    ap.add_argument("--weights_json", type=str, default="",
                    help="Đường dẫn file JSON weights tùy chỉnh (ghi đè DEFAULT_WEIGHTS)")
    args = ap.parse_args()

    device = torch.device(args.device if torch.cuda.is_available() else "cpu")

    img_size_arg = None if args.img_size <= 0 else args.img_size
    model_name_arg = None if not args.model_name else args.model_name

    ckpt_paths: List[str] = args.ckpt
    models: List[nn.Module] = []
    metas: List[Dict] = []

    for p in ckpt_paths:
        m, meta = load_checkpoint_build_model(
            ckpt_path=p,
            device=device,
            img_size_arg=img_size_arg,
            model_name_arg=model_name_arg,
        )
        models.append(m)
        metas.append(meta)
        print(f"[INFO] Loaded ckpt: {p} | model={meta.get('model_name')} | img_size={meta.get('img_size')} | best_thr={meta.get('best_thr')}")

    # phương pháp: lấy từ ckpt đầu tiên (giả định cùng schema)
    method_names = list(metas[0].get("method_names", []))
    img_size = (args.img_size if args.img_size > 0 else metas[0].get("img_size", 384))

    if args.threshold >= 0:
        thr = float(args.threshold)
    else:
        if len(metas) == 1:
            thr = float(metas[0].get("best_thr", 0.5))
        else:
            # mặc định: trung bình best_thr các model (khuyến nghị: nên tự tune thr ensemble riêng)
            thr = float(sum(m.get("best_thr", 0.5) for m in metas) / max(1, len(metas)))
    print(f"[INFO] Using threshold={thr:.6f} | num_models={len(models)}")

    # ── Resolve weights ──────────────────────────────────────────────────────
    bin_w: Optional[List[float]] = None
    met_w: Optional[Dict[str, List[float]]] = None

    weights_cfg = None
    if args.weights_json:
        import json as _json
        with open(args.weights_json, encoding="utf-8") as _f:
            weights_cfg = _json.load(_f)
        print(f"[INFO] Loaded weights from: {args.weights_json}")
    elif args.use_weights or args.use_bin_weights or args.use_met_weights:
        weights_cfg = DEFAULT_WEIGHTS
        print(f"[INFO] Using DEFAULT_WEIGHTS (built-in)")

    use_bin = args.use_weights or args.use_bin_weights
    use_met = args.use_weights or args.use_met_weights

    if weights_cfg and len(models) > 1:
        order = weights_cfg.get("model_order", [])
        if len(order) == len(models):
            if use_bin:
                bin_w = [weights_cfg["bin_weights"].get(n, 1.0) for n in order]
                print(f"[INFO] bin_weights ON : {dict(zip(order, [round(w,4) for w in bin_w]))}")
            else:
                print(f"[INFO] bin_weights OFF (simple average)")
            if use_met:
                met_w_raw = weights_cfg.get("met_weights", {})
                met_w = {m: [met_w_raw[m].get(n, 1.0) for n in order] for m in met_w_raw}
                print(f"[INFO] met_weights ON  for {list(met_w.keys())}")
            else:
                print(f"[INFO] met_weights OFF (simple average)")
        else:
            print(f"[WARN] model_order length ({len(order)}) != num_models ({len(models)}), weights ignored")
    elif weights_cfg and len(models) == 1:
        print(f"[INFO] Single model — weights ignored")

    retina_det = None
    if args.detector_backend == "retinaface":
        try:
            retina_det = RetinaFaceLargest(device=device, det_thresh=args.det_thr, det_size=640)
        except Exception as e:
            print(f"[WARN] RetinaFace init failed ({e}); fallback to MediaPipe.")
            args.detector_backend = "mediapipe"

    root = Path(args.root)
    real_dir = root / "real"
    fake_dir = root / "fake"

    if (not args.skip_real) and real_dir.is_dir():
        eval_real_dir(models, device, img_size, real_dir, args.out_real, thr, args.batch,
                      detector_backend=args.detector_backend, bbox_scale=args.bbox_scale,
                      fake_index=args.fake_index, retina_det=retina_det,
                      bin_weights=bin_w, limit_videos=args.limit_videos,
                      method_names=method_names)
        print(f"[DONE] Real -> {args.out_real}")
    else:
        print("[SKIP] Real evaluation skipped.")

    if (not args.skip_fake) and fake_dir.is_dir():
        eval_fake_dir(models, device, img_size, fake_dir, args.out_fake, thr, args.batch,
                      detector_backend=args.detector_backend, bbox_scale=args.bbox_scale,
                      method_names=method_names, fake_index=args.fake_index, retina_det=retina_det,
                      bin_weights=bin_w, met_weights=met_w, limit_videos=args.limit_videos)
        print(f"[DONE] Fake -> {args.out_fake}")
    else:
        print("[SKIP] Fake evaluation skipped.")

    # ── Tự động build Excel nếu có --out_xlsx ──
    out_xlsx = args.out_xlsx
    if not out_xlsx:
        # tự đặt tên theo tên file csv output
        base = os.path.splitext(args.out_fake)[0].replace("results_fake", "").replace("_fake", "").strip("_")
        suffix = ("_" + base) if base else ""
        out_xlsx = f"results{suffix}.xlsx"

    build_excel_report(
        fake_csv=args.out_fake if not args.skip_fake else "",
        real_csv=args.out_real if not args.skip_real else "",
        out_xlsx=out_xlsx,
    )

    # ── Confusion Matrix Excel ──
    if out_xlsx:
        build_confusion_matrix_excel(
            fake_csv=args.out_fake if not args.skip_fake else "",
            real_csv=args.out_real if not args.skip_real else "",
            out_xlsx=out_xlsx,
            model_name=args.cm_model_name,
        )


if __name__ == "__main__":
    main()