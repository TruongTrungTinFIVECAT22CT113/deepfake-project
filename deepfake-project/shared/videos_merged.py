import subprocess
import os
import random
import argparse
import cv2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ap = argparse.ArgumentParser()
ap.add_argument("--resolution", type=str, required=True,
                help="VD: 640x480 | 960x720 | 1024x768")
args = ap.parse_args()

try:
    w, h = args.resolution.lower().split("x")
    w, h = int(w), int(h)
except:
    raise ValueError(f"--resolution phải có dạng WxH, ví dụ: 640x480. Nhận được: {args.resolution}")

# ── Cấu hình đường dẫn ────────────────────────────────────────────────────────
base_fake  = rf"H:\deepfake-project\deepfake-project\data\videos_test\{args.resolution}\fake"
base_real  = rf"H:\deepfake-project\deepfake-project\data\videos_test\{args.resolution}\real\Faceforensics"
output_dir = rf"H:\deepfake-project\deepfake-project\data\videos_merged\{args.resolution}"
report_path = os.path.join(output_dir, "report.xlsx")

print(f"Resolution : {w}x{h}")
print(f"Output     : {output_dir}")

fake_folders = [
    "Audio2Animation",
    "Deepfakes",
    "Face2Face",
    "FaceShifter",
    "FaceSwap",
    "NeuralTextures",
    "Video2VideoID"
]

LABEL_COLORS = {
    "Audio2Animation":  "D7CCC8",
    "Deepfakes":        "FFCDD2",
    "Face2Face":        "FFF59D",
    "FaceShifter":      "C8E6C9",
    "FaceSwap":         "B3E5FC",
    "NeuralTextures":   "E1BEE7",
    "Video2VideoID":    "FFE0B2",
    "Fake_Unknown":     "E0E0E0",
    "Real":             "FFFFFF",
}

os.makedirs(output_dir, exist_ok=True)

def get_frame_count(video_path):
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        return 0
    count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    cap.release()
    return count

# ── Tạo workbook Excel ────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws_detail  = wb.active
ws_detail.title = "Chi tiết"
ws_summary = wb.create_sheet("Tổng hợp")

header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill  = PatternFill("solid", start_color="2E4057")
center_align = Alignment(horizontal="center", vertical="center")
left_align   = Alignment(horizontal="left",   vertical="center")
thin_border  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

detail_headers = ["Video #", "Đoạn", "Nhãn", "Loại", "Đường dẫn", "Số Frames", "Thứ tự trong video ghép"]
ws_detail.append(detail_headers)
for col, _ in enumerate(detail_headers, 1):
    cell = ws_detail.cell(row=1, column=col)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = center_align; cell.border = thin_border
ws_detail.row_dimensions[1].height = 22
ws_detail.freeze_panes = "A2"

summary_headers = ["Video #", "Tổng Frames"] + \
    [f"Frames_{lbl}" for lbl in fake_folders + ["Real"]] + ["Thứ tự ghép (label)"]
ws_summary.append(summary_headers)
for col, _ in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = center_align; cell.border = thin_border
ws_summary.row_dimensions[1].height = 22
ws_summary.freeze_panes = "A2"

# ── Xử lý từng video ─────────────────────────────────────────────────────────
success    = 0
failed     = []
detail_row = 2

for i in range(900, 1000):
    filename    = f"{i}.mp4"
    video_paths = []
    missing     = False

    for folder in fake_folders:
        path = os.path.join(base_fake, folder, filename)
        if os.path.exists(path):
            video_paths.append((folder, path, "Fake"))
        else:
            print(f"[WARN] Thiếu: {path}")
            missing = True
            break

    if missing:
        failed.append(i); continue

    real_path = os.path.join(base_real, filename)
    if os.path.exists(real_path):
        video_paths.append(("Real", real_path, "Real"))
    else:
        print(f"[WARN] Thiếu real: {real_path}")
        failed.append(i); continue

    random.shuffle(video_paths)

    temp_list = os.path.join(output_dir, f"_temp_{i}.txt")
    with open(temp_list, "w", encoding="utf-8", newline="\n") as f:
        for _, path, _ in video_paths:
            f.write(f"file '{path}'\n")

    output_file = os.path.join(output_dir, filename)
    cmd = ["ffmpeg", "-f", "concat", "-safe", "0",
           "-i", temp_list,
           "-vf", f"scale={w}:{h},setsar=1",
           "-c:v", "h264_nvenc",
           "-preset", "p4",
           "-tune", "hq",
           "-rc", "vbr",
           "-cq", "18",
           "-b:v", "0",
           "-c:a", "aac",
           output_file, "-y"]

    print(f"Ghép video {i} ({args.resolution})...")
    result = subprocess.run(cmd, capture_output=True, text=True)
    os.remove(temp_list)

    if result.returncode != 0:
        failed.append(i)
        print(f"  ✗ Lỗi {i}:\n{result.stderr[-300:]}")
        continue

    success += 1

    frames_by_label = {}
    for order_idx, (label, path, kind) in enumerate(video_paths, 1):
        frames = get_frame_count(path)
        frames_by_label[label] = frames

        row_data = [i, order_idx, label, kind, path, frames, order_idx]
        ws_detail.append(row_data)

        color   = LABEL_COLORS.get(label, "FFFFFF")
        fill    = PatternFill("solid", start_color=color)
        no_fill = PatternFill(fill_type=None)
        for col in range(1, len(row_data) + 1):
            cell           = ws_detail.cell(row=detail_row, column=col)
            cell.fill      = fill if col == 3 else no_fill
            cell.border    = thin_border
            cell.alignment = center_align if col != 5 else left_align
            cell.font      = Font(name="Arial", size=10)
        detail_row += 1

    total_frames = sum(frames_by_label.values())
    order_str    = " → ".join([lbl for lbl, _, _ in video_paths])
    summary_row  = [i, total_frames]
    for lbl in fake_folders + ["Real"]:
        summary_row.append(frames_by_label.get(lbl, 0))
    summary_row.append(order_str)

    ws_summary.append(summary_row)
    sum_row_idx = success + 1
    for col in range(1, len(summary_row) + 1):
        cell           = ws_summary.cell(row=sum_row_idx, column=col)
        cell.border    = thin_border
        cell.alignment = center_align if col != len(summary_row) else left_align
        cell.font      = Font(name="Arial", size=10, bold=(col == 2))

# ── Độ rộng cột ───────────────────────────────────────────────────────────────
for idx, w_ in enumerate([10,8,20,8,70,12,20], 1):
    ws_detail.column_dimensions[get_column_letter(idx)].width = w_
for idx, w_ in enumerate([10,14]+[16]*8+[80], 1):
    ws_summary.column_dimensions[get_column_letter(idx)].width = w_

# ── Sheet Chú thích ───────────────────────────────────────────────────────────
ws_legend = wb.create_sheet("Chú thích")
for col, val in enumerate(["Nhãn","Loại","Màu"], 1):
    cell = ws_legend.cell(row=1, column=col)
    cell.value = val; cell.font = header_font; cell.fill = header_fill
    cell.alignment = center_align; cell.border = thin_border

for r, (lbl, kind) in enumerate([(l,"Fake") for l in fake_folders]+[("Real","Real")], 2):
    color = LABEL_COLORS[lbl]
    fill  = PatternFill("solid", start_color=color)
    for col, val in enumerate([lbl, kind, f"#{color}"], 1):
        cell = ws_legend.cell(row=r, column=col)
        cell.value = val; cell.fill = fill; cell.border = thin_border
        cell.alignment = center_align; cell.font = Font(name="Arial", size=10)

ws_legend.column_dimensions["A"].width = 22
ws_legend.column_dimensions["B"].width = 10
ws_legend.column_dimensions["C"].width = 14

wb.save(report_path)

print(f"\n{'='*55}")
print(f"✓ Hoàn thành: {success}/100 videos ghép thành công")
print(f"✓ Báo cáo Excel: {report_path}")
if failed:
    print(f"✗ Thất bại ({len(failed)}): {failed}")